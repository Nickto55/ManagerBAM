
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd

class ExcelSaver:

    
    def __init__(self, filename="result_data.xlsx"):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        
        # Стили
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        self.code_font = Font(name="Consolas", size=10)
        self.code_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        self.wrap_alignment = Alignment(wrap_text=True, vertical="top")

    def _create_workbook(self, sheet_name=None):
        """Загружает существующий файл или создает новую книгу Excel."""
        output_path = os.path.join(os.getcwd(), self.filename)

        if os.path.exists(output_path):
            self.workbook = openpyxl.load_workbook(output_path)
        else:
            self.workbook = openpyxl.Workbook()
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])

        # 🔧 Создаём или получаем нужный лист
        if sheet_name and sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
            # Опционально: очистить старый лист
            self.sheet.delete_rows(1, self.sheet.max_row)
        elif sheet_name:
            self.sheet = self.workbook.create_sheet(sheet_name)
        else:
            self.sheet = self.workbook.active
            self.sheet.title = "Что-то пошло не так"

    def _auto_resize_columns(self):
        """Автоматически подбирает ширину столбцов."""
        for column in self.sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                if column_letter in ['E','F']: max_length=2
                if column_letter in ['B','G','I',]: max_length=7
            adjusted_width = min(max_length + 4, 80)
            self.sheet.column_dimensions[column_letter].width = adjusted_width
            
    def _auto_resize_rows(self):
        """Автоматически подбирает высоту строк для многострочного текста."""
        for row in self.sheet.iter_rows(min_row=2):
            max_lines = 1
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                        lines = cell.value.count('\n') + 1
                        max_lines = max(max_lines, lines)

            self.sheet.row_dimensions[row[0].row].height = max(30, max_lines * 15)

    
    def save(self, data_dict, sheet_name=None):
        if not data_dict:
            raise ValueError("Передан пустой словарь данных")
            
        self._create_workbook(sheet_name)
        
        if sheet_name:
            self.sheet.title = sheet_name
            
        # Получаем все уникальные ключи из внутренних словарей для заголовков
        headers = ["Дсе","Уп","Имя изделия","Наименование","Рц","Рц из Сз","Дата из письма","Инф из письма","Подписано","Комментарии","№Жп","Дсе ЖП", "Дата создания",'Комментарий']
        
        # Добавляем колонку с именем теста/функции
        headers = headers
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = self.sheet.cell(1, col_idx, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
            
        # Записываем данные
        row_idx = 2
        name_idxs = 0
        last_name = "Дсе_+_"
        for name, inner_dict in data_dict.items():
            if last_name[:last_name.index("_+_")] != name[:name.index("_+_")]:
                name_idxs+=1
            for rc_s, row_from_rc in inner_dict.items():
                for col_idx, header in enumerate(headers, 1):
                    if header == "Дсе":
                        value = name[:name.index("_+_")]
                    else:
                        if isinstance(row_from_rc, str): continue
                        value = row_from_rc.get(header, '')

                    if not isinstance(row_from_rc, str):
                        cell = self.sheet.cell(row_idx, col_idx, f"{value}")
                        cell.border = self.border
                        cell.alignment = self.wrap_alignment

                        if name_idxs % 2 != 0:
                            cell.fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
                        # if header in ['Дсе', 'input', 'output', 'expected']:

                            # cell.font = self.code_font
                            # cell.fill = self.code_fill
                            # cell.border = self.border

                if not isinstance(row_from_rc, str):
                    row_idx += 1
            last_name = name

            
        # Авторазмер колонок и строк
        self._auto_resize_columns()
        self._auto_resize_rows()
        
        # Замораживаем заголовок
        self.sheet.freeze_panes = 'A2'
        
        # Сохраняем
        output_path = f"{os.getcwd()}/{self.filename}"
        self.workbook.save(output_path)
        return output_path




